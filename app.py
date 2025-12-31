"""
================================================================================
REVFINDER AI - Aplicação Web Streamlit
================================================================================

Interface web profissional para análise de notas fiscais e identificação
de PIS/COFINS pagos indevidamente em produtos monofásicos.

FUNCIONALIDADES:
----------------
- Upload múltiplo de arquivos XML de NF-e
- Análise automática com IA e cache inteligente
- Dashboard com métricas em tempo real
- Download de relatório Excel
- Visualização detalhada dos erros encontrados

COMO EXECUTAR:
--------------
    streamlit run app.py

Autor: Grande Mestre
Versão: 1.0
Data: Dezembro/2025
================================================================================
"""

import streamlit as st
import pandas as pd
import tempfile
import os
import sys
from datetime import datetime
from io import BytesIO

# Adiciona o diretório src ao path para imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Imports dos módulos do RevFinder (sem IA - lazy loading)
from src.core.parser import NFeParser
from src.core.ncm_database import NCMDatabase
from src.utils.exporter import ReportGenerator
# FiscalAuditorAgent é importado sob demanda em initialize_ai_agent()

# =============================================================================
# CONFIGURAÇÃO DA PÁGINA
# =============================================================================

st.set_page_config(
    page_title="RevFinder AI - Recuperação Tributária",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# AUTENTICAÇÃO
# =============================================================================

def check_password():
    """Verifica se a senha está correta."""
    
    def password_entered():
        """Checa a senha quando o usuário submete."""
        # Pega a senha dos secrets (Streamlit Cloud) ou .env local
        correct_password = st.secrets.get("APP_PASSWORD", os.environ.get("APP_PASSWORD", "admin123"))
        
        if st.session_state["password"] == correct_password:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Remove a senha da memória
        else:
            st.session_state["password_correct"] = False

    # Primeira execução ou não logado
    if "password_correct" not in st.session_state:
        st.markdown("""
        <div style="display: flex; justify-content: center; align-items: center; height: 60vh;">
            <div style="text-align: center; padding: 40px; background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%); 
                        border-radius: 20px; box-shadow: 0 10px 40px rgba(0,0,0,0.3);">
                <h1 style="color: white; margin-bottom: 10px;">🚀 RevFinder AI</h1>
                <p style="color: #a0c4e8; margin-bottom: 30px;">Sistema de Recuperação Tributária</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            st.text_input(
                "🔐 Digite a senha de acesso:",
                type="password",
                on_change=password_entered,
                key="password"
            )
            st.caption("Entre em contato para obter acesso.")
        return False
    
    # Senha errada
    elif not st.session_state["password_correct"]:
        st.markdown("""
        <div style="display: flex; justify-content: center; align-items: center; height: 60vh;">
            <div style="text-align: center; padding: 40px; background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%); 
                        border-radius: 20px; box-shadow: 0 10px 40px rgba(0,0,0,0.3);">
                <h1 style="color: white; margin-bottom: 10px;">🚀 RevFinder AI</h1>
                <p style="color: #a0c4e8; margin-bottom: 30px;">Sistema de Recuperação Tributária</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            st.text_input(
                "🔐 Digite a senha de acesso:",
                type="password",
                on_change=password_entered,
                key="password"
            )
            st.error("❌ Senha incorreta. Tente novamente.")
        return False
    
    # Senha correta
    else:
        return True

# =============================================================================
# CSS CUSTOMIZADO
# =============================================================================

st.markdown("""
<style>
    /* Cards de métricas */
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px;
        border-radius: 10px;
        color: white;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    
    .metric-card h2 {
        margin: 0;
        font-size: 2.5rem;
        font-weight: bold;
    }
    
    .metric-card p {
        margin: 5px 0 0 0;
        opacity: 0.9;
    }
    
    /* Header */
    .main-header {
        background: linear-gradient(90deg, #1a1a2e 0%, #16213e 100%);
        padding: 20px;
        border-radius: 10px;
        margin-bottom: 20px;
        color: white;
    }
    
    /* Tabela de resultados */
    .dataframe {
        font-size: 14px;
    }
    
    /* Botão de análise */
    .stButton > button {
        width: 100%;
        background: linear-gradient(90deg, #11998e 0%, #38ef7d 100%);
        color: white;
        font-weight: bold;
        padding: 15px;
        font-size: 18px;
        border: none;
        border-radius: 10px;
    }
    
    .stButton > button:hover {
        background: linear-gradient(90deg, #38ef7d 0%, #11998e 100%);
    }
    
    /* Upload area */
    .uploadedFile {
        background-color: #f0f2f6;
        border-radius: 5px;
        padding: 10px;
    }
    
    /* Success/Error messages */
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
    
    .error-box {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        color: #721c24;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# FUNÇÕES AUXILIARES
# =============================================================================

@st.cache_resource
def load_database():
    """Carrega o banco de dados de NCMs (com cache do Streamlit)."""
    db_path = os.path.join(os.path.dirname(__file__), "src", "database", "ncm_rules.json")
    return NCMDatabase(db_path)


def load_ai_agent():
    """
    Carrega o agente de IA com lazy loading.
    
    Só inicializa quando realmente precisar, evitando warnings
    do CrewAI no startup do Streamlit.
    """
    # Usa session_state para cache manual (evita reload)
    if 'ia_agent' not in st.session_state:
        st.session_state.ia_agent = None
        st.session_state.ia_checked = False
    
    return st.session_state.ia_agent


def initialize_ai_agent():
    """Inicializa o agente de IA (chamado só quando necessário)."""
    if st.session_state.get('ia_agent') is None and not st.session_state.get('ia_checked'):
        try:
            import warnings
            import logging
            
            # Silencia TODOS os warnings e logs do CrewAI
            warnings.filterwarnings("ignore")
            logging.getLogger("crewai").setLevel(logging.CRITICAL)
            logging.getLogger("langchain").setLevel(logging.CRITICAL)
            
            # Suprime o erro de signal (o problema principal)
            import signal
            original_signal = signal.signal
            def dummy_signal(*args, **kwargs):
                return None
            signal.signal = dummy_signal
            
            try:
                # LAZY IMPORT - só importa quando precisa
                from src.agents.auditor import FiscalAuditorAgent
                st.session_state.ia_agent = FiscalAuditorAgent()
            finally:
                # Restaura o signal original
                signal.signal = original_signal
            
            st.session_state.ia_checked = True
        except Exception as e:
            st.session_state.ia_agent = None
            st.session_state.ia_checked = True
    
    return st.session_state.ia_agent


def analyze_item(item: dict, ncm_db: NCMDatabase, stats: dict) -> dict | None:
    """
    Analisa um item da nota fiscal.
    
    Replica a lógica do main.py para uso no Streamlit.
    A IA só é carregada se realmente precisar (lazy loading).
    """
    # Se não pagou imposto, não tem o que recuperar
    if item.get('imposto_total', 0) <= 0:
        return None
    
    # Verificação pelo Banco de Dados (NCM + Keywords + Cache)
    resultado_db = ncm_db.verificar_item(item['ncm'], item['produto'])
    
    if resultado_db['is_monofasico']:
        # Determina a fonte
        if resultado_db['fonte'] == 'banco_dados':
            stats['banco_dados'] = stats.get('banco_dados', 0) + 1
            origem = "Banco de Dados"
            motivo = f"NCM {item['ncm']} é monofásico - {resultado_db['descricao']}"
        elif resultado_db['fonte'] == 'cache_ia':
            stats['cache_ia'] = stats.get('cache_ia', 0) + 1
            stats['ia_economizada'] = stats.get('ia_economizada', 0) + 1
            origem = "Cache IA"
            motivo = f"Aprendizado anterior - {resultado_db['descricao']}"
        else:
            stats['keywords'] = stats.get('keywords', 0) + 1
            stats['ia_economizada'] = stats.get('ia_economizada', 0) + 1
            origem = "Keywords"
            keyword = resultado_db.get('keyword_encontrada', '')
            motivo = f"Keyword '{keyword}' - {resultado_db['descricao']}"
        
        return {
            "chave_acesso": item.get('chave_acesso', ''),
            "numero_nota": item.get('numero_nota', ''),
            "data_emissao": item.get('data_emissao', ''),
            "cnpj_emitente": item.get('cnpj_emitente', ''),
            "nome_emitente": item.get('nome_emitente', ''),
            "produto": item['produto'],
            "ncm": item['ncm'],
            "ncm_correto": resultado_db['ncm_correto'],
            "imposto_recuperavel": item['imposto_total'],
            "motivo": motivo,
            "origem_analise": origem,
            "base_legal": resultado_db['base_legal'],
            "confianca": resultado_db.get('confianca', 'alta')
        }
    
    # Se já tem no cache (não monofásico), não chama IA
    if resultado_db.get('fonte') == 'cache_ia':
        stats['ia_economizada'] = stats.get('ia_economizada', 0) + 1
        return None
    
    # Consulta IA (último recurso) - LAZY LOADING
    # Só carrega a IA agora, quando realmente precisa
    ia_auditor = initialize_ai_agent()
    
    if ia_auditor:
        resultado_ia = ia_auditor.analyze_item(
            descricao=item['produto'],
            ncm_errado=item['ncm'],
            valor_item=item['valor_total']
        )
        
        # Salva aprendizado
        ncm_db.salvar_aprendizado_ia(
            nome_produto=item['produto'],
            is_monofasico=(resultado_ia[0] == True),
            ncm_sugerido=resultado_ia[1],
            motivo=resultado_ia[2]
        )
        
        if resultado_ia[0] == True:
            stats['ia'] = stats.get('ia', 0) + 1
            
            return {
                "chave_acesso": item.get('chave_acesso', ''),
                "numero_nota": item.get('numero_nota', ''),
                "data_emissao": item.get('data_emissao', ''),
                "cnpj_emitente": item.get('cnpj_emitente', ''),
                "nome_emitente": item.get('nome_emitente', ''),
                "produto": item['produto'],
                "ncm": item['ncm'],
                "ncm_correto": resultado_ia[1],
                "imposto_recuperavel": item['imposto_total'],
                "motivo": resultado_ia[2],
                "origem_analise": "Agente IA",
                "base_legal": ncm_db.get_base_legal(),
                "confianca": "media"
            }
    
    return None


def process_xml_files(uploaded_files, ncm_db: NCMDatabase) -> tuple:
    """
    Processa múltiplos arquivos XML.
    
    A IA só é carregada se realmente precisar (lazy loading).
    
    Returns:
        tuple: (lista_erros, stats, total_itens, total_notas)
    """
    parser = NFeParser()
    erros_encontrados = []
    stats = {
        'banco_dados': 0,
        'keywords': 0,
        'cache_ia': 0,
        'ia': 0,
        'ia_economizada': 0
    }
    total_itens = 0
    total_notas = 0
    ia_auditor = None  # Lazy loading - só carrega se precisar
    
    # Barra de progresso
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, uploaded_file in enumerate(uploaded_files):
        # Atualiza progresso
        progress = (idx + 1) / len(uploaded_files)
        progress_bar.progress(progress)
        status_text.text(f"📂 Processando: {uploaded_file.name}...")
        
        # Salva arquivo temporário
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xml') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_path = tmp_file.name
        
        try:
            # Parseia XML
            itens_nota = parser.parse(tmp_path)
            
            if itens_nota:
                total_notas += 1
                total_itens += len(itens_nota)
                
                # Analisa cada item
                for item in itens_nota:
                    erro = analyze_item(item, ncm_db, stats)
                    if erro:
                        erros_encontrados.append(erro)
        finally:
            # Remove arquivo temporário
            os.unlink(tmp_path)
    
    # Limpa barra de progresso
    progress_bar.empty()
    status_text.empty()
    
    return erros_encontrados, stats, total_itens, total_notas


def create_excel_download(erros: list, total_recuperavel: float) -> bytes:
    """
    Cria arquivo Excel para download com disclaimer.
    
    Args:
        erros: Lista de erros encontrados
        total_recuperavel: Valor total recuperável
    
    Returns:
        bytes: Arquivo Excel em bytes
    """
    if not erros:
        return None
    
    # Importa openpyxl para formatação avançada
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    df = pd.DataFrame(erros)
    
    # Renomeia colunas
    column_mapping = {
        "chave_acesso": "Chave de Acesso",
        "numero_nota": "Nº Nota",
        "data_emissao": "Data Emissão",
        "cnpj_emitente": "CNPJ Emitente",
        "nome_emitente": "Emitente",
        "produto": "Produto",
        "ncm": "NCM Atual",
        "ncm_correto": "NCM Correto",
        "imposto_recuperavel": "Valor (R$)",
        "motivo": "Motivo",
        "origem_analise": "Fonte",
        "base_legal": "Base Legal"
    }
    
    df = df.rename(columns=column_mapping)
    
    # Cria Excel em memória
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Aba principal com dados
        df.to_excel(writer, index=False, sheet_name='Produtos Monofasicos', startrow=4)
        
        # Acessa a planilha para formatação
        ws = writer.sheets['Produtos Monofasicos']
        
        # --- CABEÇALHO ---
        ws.merge_cells('A1:L1')
        ws['A1'] = '🚀 REVFINDER AI - Relatório de Recuperação Tributária'
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:L2')
        ws['A2'] = f'Total Potencial de Recuperação: R$ {total_recuperavel:,.2f}'
        ws['A2'].font = Font(bold=True, size=14, color='008000')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:L3')
        ws['A3'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")} | {len(erros)} produtos identificados'
        ws['A3'].font = Font(italic=True, size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # --- FORMATA CABEÇALHO DA TABELA ---
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[5]:  # Linha 5 é o cabeçalho (startrow=4 + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # --- AJUSTA LARGURA DAS COLUNAS ---
        ws.column_dimensions['A'].width = 50  # Chave de Acesso
        ws.column_dimensions['B'].width = 12  # Nº Nota
        ws.column_dimensions['C'].width = 12  # Data
        ws.column_dimensions['D'].width = 18  # CNPJ
        ws.column_dimensions['E'].width = 25  # Emitente
        ws.column_dimensions['F'].width = 40  # Produto
        ws.column_dimensions['G'].width = 12  # NCM Atual
        ws.column_dimensions['H'].width = 12  # NCM Correto
        ws.column_dimensions['I'].width = 12  # Valor
        ws.column_dimensions['J'].width = 30  # Motivo
        ws.column_dimensions['K'].width = 15  # Fonte
        ws.column_dimensions['L'].width = 35  # Base Legal
        
        # --- ABA DE DISCLAIMER ---
        ws_disclaimer = writer.book.create_sheet('⚠️ IMPORTANTE')
        
        disclaimer_texts = [
            ('⚠️ AVISO IMPORTANTE', True, 16),
            ('', False, 10),
            ('Este relatório identifica produtos com tributação MONOFÁSICA de PIS/COFINS.', False, 11),
            ('O valor apresentado representa o POTENCIAL de recuperação.', False, 11),
            ('', False, 10),
            ('📋 ANTES DE SOLICITAR RESTITUIÇÃO, VERIFIQUE:', True, 12),
            ('', False, 10),
            ('1. CONSULTE SEU CONTADOR', True, 11),
            ('   Pergunte: "Você está segregando as receitas de produtos monofásicos no PGDAS-D?"', False, 10),
            ('', False, 10),
            ('2. SE O CONTADOR JÁ SEGREGA:', True, 11),
            ('   ✅ Sua empresa já está pagando corretamente', False, 10),
            ('   ✅ NÃO há valores a recuperar', False, 10),
            ('   ✅ Este relatório serve apenas como conferência', False, 10),
            ('', False, 10),
            ('3. SE O CONTADOR NÃO SEGREGA:', True, 11),
            ('   💰 Há valores a recuperar dos últimos 5 anos', False, 10),
            ('   📝 Solicite a retificação das declarações', False, 10),
            ('   🏦 Peça restituição ou compensação via PER/DCOMP', False, 10),
            ('', False, 10),
            ('📜 BASE LEGAL:', True, 12),
            ('Lei nº 13.097/2015, arts. 14 a 36', False, 10),
            ('Decreto nº 8.442/2015', False, 10),
            ('Solução de Consulta COSIT nº 99002/2024', False, 10),
            ('', False, 10),
            ('⚡ PRODUTOS MONOFÁSICOS (alíquota zero no Simples):', True, 12),
            ('• Água mineral e gaseificada', False, 10),
            ('• Refrigerantes', False, 10),
            ('• Cervejas (com e sem álcool)', False, 10),
            ('• Energéticos e isotônicos', False, 10),
            ('• Chás e sucos prontos para consumo', False, 10),
            ('', False, 10),
            ('❌ NÃO SÃO MONOFÁSICOS:', True, 12),
            ('• Vinhos e espumantes', False, 10),
            ('• Destilados (whisky, vodka, gin, etc.)', False, 10),
            ('• Drinks preparados no estabelecimento', False, 10),
            ('• Café', False, 10),
            ('', False, 10),
            ('---', False, 10),
            ('Relatório gerado por RevFinder AI v2.3', False, 9),
            ('Desenvolvido por Grande Mestre - 2025', False, 9),
        ]
        
        for idx, (text, is_bold, size) in enumerate(disclaimer_texts, start=1):
            cell = ws_disclaimer.cell(row=idx, column=1, value=text)
            cell.font = Font(bold=is_bold, size=size)
            if is_bold and size >= 12:
                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        ws_disclaimer.column_dimensions['A'].width = 80
        
        # --- ABA DE RESUMO ---
        ws_resumo = writer.book.create_sheet('📊 Resumo')
        
        resumo_data = [
            ('RESUMO DA ANÁLISE', '', ''),
            ('', '', ''),
            ('Total de Produtos Analisados:', len(erros), ''),
            ('Valor Total Potencial:', f'R$ {total_recuperavel:,.2f}', ''),
            ('', '', ''),
            ('Por Fonte de Identificação:', '', ''),
        ]
        
        # Conta por fonte
        fonte_count = df['Fonte'].value_counts().to_dict() if 'Fonte' in df.columns else {}
        for fonte, count in fonte_count.items():
            resumo_data.append((f'  • {fonte}', count, 'produtos'))
        
        for idx, (col1, col2, col3) in enumerate(resumo_data, start=1):
            ws_resumo.cell(row=idx, column=1, value=col1)
            ws_resumo.cell(row=idx, column=2, value=col2)
            ws_resumo.cell(row=idx, column=3, value=col3)
            if idx == 1:
                ws_resumo.cell(row=idx, column=1).font = Font(bold=True, size=14)
        
        ws_resumo.column_dimensions['A'].width = 35
        ws_resumo.column_dimensions['B'].width = 20
    
    return output.getvalue()


# =============================================================================
# INTERFACE PRINCIPAL
# =============================================================================

def main():
    """Função principal da aplicação."""
    
    # -----------------------------------------------------------------
    # VERIFICAÇÃO DE SENHA
    # -----------------------------------------------------------------
    if not check_password():
        return  # Para aqui se não estiver autenticado
    
    # -----------------------------------------------------------------
    # HEADER
    # -----------------------------------------------------------------
    st.markdown("""
    <div class="main-header">
        <h1>🚀 RevFinder AI</h1>
        <p>Sistema Inteligente de Recuperação Tributária - PIS/COFINS Monofásico</p>
    </div>
    """, unsafe_allow_html=True)
    
    # -----------------------------------------------------------------
    # SIDEBAR
    # -----------------------------------------------------------------
    with st.sidebar:
        st.header("⚙️ Configurações")
        
        # Carrega banco de dados
        ncm_db = load_database()
        db_stats = ncm_db.get_estatisticas()
        
        st.success(f"✅ Base carregada")
        st.caption(f"📊 {db_stats['total_ncms']} NCMs | {db_stats['total_keywords']} keywords")
        
        # Status do cache
        cache_stats = ncm_db.get_estatisticas_cache()
        if cache_stats['total_produtos'] > 0:
            st.info(f"🧠 Cache: {cache_stats['total_produtos']} produtos aprendidos")
        
        st.divider()
        
        # Status da IA (lazy - não carrega ainda)
        st.info("🤖 IA: Disponível sob demanda")
        
        st.divider()
        
        # Informações
        st.header("📜 Base Legal")
        st.caption(db_stats['base_legal'])
        
        st.divider()
        
        st.header("ℹ️ Sobre")
        st.caption("""
        **RevFinder AI** analisa notas fiscais
        eletrônicas (NF-e) para identificar
        PIS/COFINS pagos indevidamente em
        produtos com tributação monofásica.
        
        Desenvolvido por **Grande Mestre**
        """)
        
        st.divider()
        
        # Botão de Logout
        if st.button("🚪 Sair", type="secondary"):
            st.session_state["password_correct"] = False
            st.rerun()
    
    # -----------------------------------------------------------------
    # ÁREA PRINCIPAL
    # -----------------------------------------------------------------
    
    # Upload de arquivos
    st.header("📁 Upload de Notas Fiscais")
    
    uploaded_files = st.file_uploader(
        "Arraste seus arquivos XML ou clique para selecionar",
        type=['xml'],
        accept_multiple_files=True,
        help="Selecione um ou mais arquivos XML de NF-e para análise"
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} arquivo(s) carregado(s)")
        
        # Lista arquivos
        with st.expander("📋 Ver arquivos carregados"):
            for f in uploaded_files:
                st.text(f"  • {f.name}")
    
    st.divider()
    
    # Botão de análise
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        analyze_button = st.button(
            "🔍 ANALISAR NOTAS FISCAIS",
            type="primary",
            disabled=not uploaded_files
        )
    
    # -----------------------------------------------------------------
    # PROCESSAMENTO E RESULTADOS
    # -----------------------------------------------------------------
    
    if analyze_button and uploaded_files:
        
        with st.spinner("🔄 Analisando notas fiscais..."):
            erros, stats, total_itens, total_notas = process_xml_files(
                uploaded_files, ncm_db
            )
        
        total_recuperavel = sum(e['imposto_recuperavel'] for e in erros)
        
        # -----------------------------------------------------------------
        # MÉTRICAS
        # -----------------------------------------------------------------
        st.header("📊 Resultado da Análise")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                label="💰 Total Recuperável",
                value=f"R$ {total_recuperavel:.2f}"
            )
        
        with col2:
            st.metric(
                label="📄 Notas Analisadas",
                value=total_notas
            )
        
        with col3:
            st.metric(
                label="🚨 Erros Encontrados",
                value=len(erros)
            )
        
        with col4:
            total_consultas = stats['banco_dados'] + stats['keywords'] + stats['cache_ia'] + stats['ia']
            economia = (stats['ia_economizada'] / max(total_consultas, 1)) * 100
            st.metric(
                label="🧠 Economia IA",
                value=f"{economia:.0f}%"
            )
        
        st.divider()
        
        # -----------------------------------------------------------------
        # ESTATÍSTICAS DETALHADAS
        # -----------------------------------------------------------------
        st.subheader("📈 Estatísticas de Identificação")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.info(f"🗄️ Banco de Dados: **{stats['banco_dados']}**")
        
        with col2:
            st.info(f"🔤 Keywords: **{stats['keywords']}**")
        
        with col3:
            st.info(f"🧠 Cache IA: **{stats['cache_ia']}**")
        
        with col4:
            st.info(f"🤖 IA Nova: **{stats['ia']}**")
        
        st.divider()
        
        # -----------------------------------------------------------------
        # TABELA DE RESULTADOS
        # -----------------------------------------------------------------
        if erros:
            st.subheader("📋 Detalhamento dos Erros")
            
            # Cria DataFrame para exibição
            df_display = pd.DataFrame(erros)
            
            # Seleciona e renomeia colunas para exibição
            columns_display = {
                'produto': 'Produto',
                'ncm': 'NCM Atual',
                'ncm_correto': 'NCM Correto',
                'imposto_recuperavel': 'Valor (R$)',
                'origem_analise': 'Fonte',
                'numero_nota': 'Nota'
            }
            
            df_show = df_display[[c for c in columns_display.keys() if c in df_display.columns]]
            df_show = df_show.rename(columns=columns_display)
            
            # Formata valores
            if 'Valor (R$)' in df_show.columns:
                df_show['Valor (R$)'] = df_show['Valor (R$)'].apply(lambda x: f"R$ {x:.2f}")
            
            st.dataframe(
                df_show,
                hide_index=True
            )
            
            st.divider()
            
            # -----------------------------------------------------------------
            # DOWNLOAD DO RELATÓRIO
            # -----------------------------------------------------------------
            st.subheader("📥 Exportar Relatório")
            
            excel_data = create_excel_download(erros, total_recuperavel)
            
            if excel_data:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Relatorio_Recuperacao_{timestamp}.xlsx"
                
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    st.download_button(
                        label="📥 BAIXAR RELATÓRIO EXCEL",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                st.success(f"✅ Relatório pronto! {len(erros)} registros | Total: R$ {total_recuperavel:.2f}")
                
                # -----------------------------------------------------------------
                # DISCLAIMER NA INTERFACE
                # -----------------------------------------------------------------
                st.divider()
                
                with st.expander("⚠️ AVISO IMPORTANTE - Leia antes de solicitar restituição", expanded=False):
                    st.markdown("""
                    ### 📋 Este relatório identifica o **POTENCIAL** de recuperação
                    
                    O valor apresentado representa produtos com tributação **monofásica** de PIS/COFINS.
                    
                    **Antes de solicitar restituição, verifique com seu contador:**
                    
                    ---
                    
                    #### ✅ Se o contador **JÁ SEGREGA** monofásicos no PGDAS-D:
                    - Sua empresa já está pagando corretamente
                    - **NÃO há valores a recuperar**
                    - Este relatório serve apenas como conferência
                    
                    ---
                    
                    #### 💰 Se o contador **NÃO SEGREGA** monofásicos:
                    - Há valores a recuperar dos últimos **5 anos**
                    - Solicite a retificação das declarações
                    - Peça restituição ou compensação via **PER/DCOMP**
                    
                    ---
                    
                    #### 📜 Base Legal:
                    - Lei nº 13.097/2015, arts. 14 a 36
                    - Decreto nº 8.442/2015
                    - Solução de Consulta COSIT nº 99002/2024
                    """)
        
        else:
            st.success("✅ Nenhum erro encontrado! Todas as notas estão corretas.")
    
    # -----------------------------------------------------------------
    # FOOTER
    # -----------------------------------------------------------------
    st.divider()
    st.caption("🚀 RevFinder AI v2.1 | Desenvolvido por Grande Mestre | 2025")


# =============================================================================
# PONTO DE ENTRADA
# =============================================================================

if __name__ == "__main__":
    main()
